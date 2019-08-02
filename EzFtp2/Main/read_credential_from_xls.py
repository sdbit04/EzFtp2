import openpyxl

workbook = openpyxl.load_workbook(r"D:\D_drive_BACKUP\Study\PycharmProjects\EzFtp2\EzFtp2\Data\credentials.xlsx")
# sheet1 = workbook.get_sheet_by_name("Sheet1")
sheet1 = workbook.active
mc = sheet1.max_column
mr = sheet1.max_row

print(mc)
print(mr)


def get_jobs_list_and_run(download_method):
    job_list = []
    for row_id in range(2,mr+1):
        # job_row = []
        job_list.append([])
        for col_id in range(1,mc+1):
            cellob = sheet1.cell(row=row_id, column=col_id)
            # print(cellob.value, end=",")
            job_list[row_id-2].append(cellob.value)

    print(job_list)

    for job in job_list:
        credential = job[1::]
        print(credential)
        download_method(*credential)




